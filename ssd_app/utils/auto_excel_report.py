
import openpyxl
from openpyxl.styles import Font, Alignment
import frappe
from frappe.utils import today, now_datetime, add_days
from ssd_app.utils.banking import export_banking_data, import_banking_data

def get_last_report_datetime():
    last_dt = frappe.db.get_single_value(
        "Auto Excel Report Settings",
        "last_report_datetime"
    )

    if not last_dt:
        last_dt = add_days(now_datetime(), -1)

    return last_dt

def update_last_report_datetime():
    frappe.db.set_single_value(
        "Auto Excel Report Settings",
        "last_report_datetime",
        now_datetime()
    )
    frappe.db.commit()

def generate_daily_banking(as_on=today()):
    last_dt_time= get_last_report_datetime()
    wb = openpyxl.Workbook()

    # -----------------------------
    # SHEET 1: Doc Received
    # -----------------------------
    data = frappe.db.sql("""
        SELECT 
            cif.inv_no AS inv_no, 
            dr.received_date AS date,
            cus.customer AS customer,
            bank.bank AS bank, 
            dr.received AS received 
        FROM `tabDoc Received` dr
        LEFT JOIN `tabCIF Sheet` cif ON cif.name = dr.inv_no
        LEFT JOIN `tabBank` bank ON bank.name = dr.bank
        LEFT JOIN `tabCustomer` cus ON cus.name = dr.customer
        WHERE dr.creation > %s
    """, (last_dt_time), as_dict=True)
    
    ws1 = wb.active
    ws1.title = "Doc Received"
    headers = ["Inv No", "Rec Date", "Customer", "Bank", "Rec Amount"]
    ws1.append(headers)

    for c in ws1[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row in data:
        ws1.append([
            row.inv_no or "",
            row.date or "",
            row.customer or "",
            row.bank or "",
            row.received or 0
        ])

    for c, w in {"A":18,"B":14,"C":25,"D":20,"E":16}.items():
        ws1.column_dimensions[c].width = w


    # -----------------------------
    # SHEET 2: Doc Negotiation
    # -----------------------------
    data2 = frappe.db.sql("""
        SELECT 
            cif.inv_no AS inv_no, 
            dn.nego_date AS date,
            cus.customer AS customer, 
            noti.notify AS notify, 
            bank.bank AS bank, 
            dn.nego_amount AS nego
        FROM `tabDoc Nego` dn
        LEFT JOIN `tabCIF Sheet` cif ON cif.name = dn.inv_no
        LEFT JOIN `tabBank` bank ON bank.name = dn.bank
        LEFT JOIN `tabCustomer` cus ON cus.name = dn.customer
        LEFT JOIN `tabNotify` noti ON noti.name = dn.notify
        WHERE dn.creation > %s 
    """, (last_dt_time), as_dict=True)


    if data2:
        ws2 = wb.create_sheet("Doc Negotiation")
        headers2 = ["Inv No", "Nego Date", "Customer", "Notify", "Bank", "Nego Amount"]
        ws2.append(headers2)

        for c in ws2[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        for row in data2:
            ws2.append([
                row.inv_no or "",
                row.date or "",
                row.customer or "",
                row.notify or "",
                row.bank or "",
                row.nego or 0
            ])

        for c, w in {"A":18,"B":14,"C":25,"D":20,"E":18,"F":18}.items():
            ws2.column_dimensions[c].width = w


    # -----------------------------
    # SHEET 3: Doc Refund
    # -----------------------------
    data3 = frappe.db.sql("""
        SELECT 
            cif.inv_no AS inv_no, 
            dr.refund_date AS date,
            cus.customer AS customer, 
            noti.notify AS notify, 
            bank.bank AS bank, 
            dr.refund_amount AS refund
        FROM `tabDoc Refund` dr
        LEFT JOIN `tabCIF Sheet` cif ON cif.name = dr.inv_no
        LEFT JOIN `tabBank` bank ON bank.name = dr.bank
        LEFT JOIN `tabCustomer` cus ON cus.name = dr.customer
        LEFT JOIN `tabNotify` noti ON noti.name = dr.notify
        WHERE dr.creation > %s 
    """, (last_dt_time), as_dict=True)


    if data3:
        ws3 = wb.create_sheet("Doc Refund")
        headers3 = ["Inv No", "Refund Date", "Customer", "Notify", "Bank", "Refund Amount"]
        ws3.append(headers3)

        for c in ws3[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        for row in data3:
            ws3.append([
                row.inv_no or "",
                row.date or "",
                row.customer or "",
                row.notify or "",
                row.bank or "",
                row.refund or 0
            ])

        for c, w in {"A":18,"B":14,"C":25,"D":20,"E":18,"F":18}.items():
            ws3.column_dimensions[c].width = w


    # -----------------------------
    # SHEET 4: CC Received  (Corrected)
    # -----------------------------
    data4 = frappe.db.sql("""
        SELECT 
            ccr.date AS date,
            cus.customer AS customer, 
            ccr.amount_usd AS cc_received
        FROM `tabCC Received` ccr
        LEFT JOIN `tabCustomer` cus ON cus.name = ccr.customer
        WHERE ccr.creation > %s 
    """, (last_dt_time), as_dict=True)

    if data4:
        ws4 = wb.create_sheet("CC Received")
        headers4 = ["Date", "Customer", "CC Amount"]
        ws4.append(headers4)

        for c in ws4[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        for row in data4:
            ws4.append([
                row.date or "",
                row.customer or "",
                row.cc_received or 0
            ])

        for c, w in {"A":18,"B":25,"C":18}.items():
            ws4.column_dimensions[c].width = w

    # -----------------------------
    # SHEET 5: Bank liability
    # -----------------------------
    imp_data = import_banking_data(as_on)
    exp_data = export_banking_data(as_on)

    if not imp_data:
        return "<p>No data found</p>"

    data5 = []

    # 1️⃣ Add import data (keep as-is)
    for row in imp_data:
        data5.append(dict(row))  # safe copy

    # 2️⃣ Add export data with renamed keys
    for row in exp_data:
        new_row = dict(row)

        # Rename keys safely, with default values
        new_row["ref_no"] = row.get("inv_no", "")
        new_row["amount_usd"] = row.get("nego", 0)

        data5.append(new_row)

    # 3️⃣ Create Excel sheet
    ws5 = wb.create_sheet("Bank Liability")
    headers5 = ["Inv No", "Bank", "Term", "Com", "Nego Amount"]
    ws5.append(headers5)

    # Make header bold & centered
    for cell in ws5[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 4️⃣ Append data rows safely
    for row in data5:
        ws5.append([
            row.get("ref_no", ""),      # safer than row["ref_no"] or ""
            row.get("bank", ""),
            row.get("p_term", ""),
            row.get("com", ""),
            row.get("amount_usd", 0)
        ])



    # -----------------------------
    # SHEET 6: Bank Liability Pivot
    # -----------------------------

    # Step 1: Build pivot structure
    pivot = {}
    p_terms = sorted({row["p_term"] for row in data5})

    for row in data5:
        bank = row["bank"]
        com = row["com"]
        term = row["p_term"]
        nego = row["amount_usd"] or 0

        pivot.setdefault(bank, {})
        pivot[bank].setdefault(com, {})
        pivot[bank][com][term] = pivot[bank][com].get(term, 0) + nego

    # Step 2: Create sheet
    ws6 = wb.create_sheet("Bank Liability Pivot")

    # Header row
    headers6 = ["Bank", "Com"] + p_terms + ["Total"]
    ws6.append(headers6)

    for c in ws6[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    # Step 3: Add pivot rows
    grand_totals = {pt: 0 for pt in p_terms}
    grand_totals["Total"] = 0

    for bank, com_data in pivot.items():
        for com, term_data in com_data.items():
            row = [bank, com]

            total = 0
            for pt in p_terms:
                val = term_data.get(pt, 0)
                row.append(val)
                total += val
                grand_totals[pt] += val

            row.append(total)
            grand_totals["Total"] += total

            ws6.append(row)

    # Step 4: Add GRAND TOTAL row
    grand_row = ["", "Grand Total"]

    for pt in p_terms:
        grand_row.append(grand_totals[pt])

    grand_row.append(grand_totals["Total"])

    ws6.append(grand_row)

    # Styling the GRAND TOTAL row
    last_row = ws6.max_row
    for cell in ws6[last_row]:
        cell.font = Font(bold=True)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0.00"

    # Step 5: Number formatting for all number cells
    for row in ws6.iter_rows(min_row=2, min_col=3):  # from column C onwards
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    # Step 6: Auto column width
    for col in ws6.columns:
        max_len = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws6.column_dimensions[column].width = max_len + 2

    # -----------------------------
    # SHEET 7: CC Received  (Corrected)
    # -----------------------------
    data7 = frappe.db.sql("""
        SELECT 
            lco.lc_open_date AS date,
            com.company_code AS com,
            bank.bank AS bank,
            lco.amount_usd AS amount,
            lco.note
        FROM `tabLC Open` lco
        LEFT JOIN `tabBank` bank ON bank.name= lco.bank
        LEFT JOIN `tabCompany` com ON com.name= lco.company
        WHERE lco.creation > %s 
    """, (last_dt_time), as_dict=True)

    if data7:
        ws7 = wb.create_sheet("LC Open")
        headers7 = ["Date", "Company", "Bank", "Amount", "Note"]
        ws7.append(headers7)

        for c in ws7[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        for row in data7:
            ws7.append([
                row.date or "",
                row.com or "",
                row.bank or "",
                row.amount or 0,
                row.note or ""
            ])

        for c, w in {"A":18,"B":25,"C":18, "D":20, "E":20}.items():
            ws7.column_dimensions[c].width = w


    # -----------------------------
    # Save workbook
    # -----------------------------
    file_path = frappe.utils.get_site_path("private", "files", "daily_banking.xlsx")
    wb.save(file_path)
    return file_path


def send_daily_banking_email():
    file_path = generate_daily_banking()

    with open(file_path, "rb") as f:
        file_data = f.read()

    file_name = f"daily_banking_{today()}.xlsx"

    frappe.sendmail(
        recipients=["ssdolui.in@gmail.com","shibsankar.dolui@uniexcelgroup.com.tw"],
        subject="Auto Excel Report",
        message="Please find the attached Excel report.",
        attachments=[{
            "fname": file_name,
            "fcontent": file_data,
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }]
    )
    update_last_report_datetime()
    return "Email sent."
